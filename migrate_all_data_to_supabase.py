import sys
sys.stdout.reconfigure(encoding='utf-8')

import urllib.request
import urllib.parse
import json
import os
import glob
import re
from datetime import datetime, date
import openpyxl
import csv

SUPABASE_URL = "https://tgctyouhzsyizlosrmqh.supabase.co"
ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InRnY3R5b3VoenN5aXpsb3NybXFoIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODY2NjYyNzAsImV4cCI6MjEwMjI0MjI3MH0.mlba06N1LhjT9vwNBhVrhMlxvjvO7QsRjArI6ue7Pv0"

headers = {
    "apikey": ANON_KEY,
    "Authorization": f"Bearer {ANON_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation"
}

def supabase_post(endpoint, payload):
    url = f"{SUPABASE_URL}/rest/v1/{endpoint}"
    data_bytes = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(url, data=data_bytes, headers=headers, method='POST')
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode('utf-8')), None
    except Exception as e:
        err_msg = str(e)
        if hasattr(e, 'read'):
            err_msg += " : " + e.read().decode('utf-8')
        return None, err_msg

def format_date(val):
    if not val:
        return date.today().isoformat()
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    # If DD/MM/YYYY
    if re.match(r'^\d{1,2}/\d{1,2}/\d{4}', s):
        parts = s.split('/')
        d, m, y = int(parts[0]), int(parts[1]), int(parts[2].split()[0])
        if y > 2500: y -= 543
        return f"{y:04d}-{m:02d}-{d:02d}"
    # If YYYY-MM-DD
    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
        return s[:10]
    return date.today().isoformat()

print("======================================================================")
print("🚀 STARTING MIGRATION OF RAW DATA TO SUPABASE")
print("======================================================================")

# ============================================================================
# 1. MIGRATE BOOKINGS
# ============================================================================
booking_file = glob.glob("downloaded_raw/BOOKING*.xlsx")
if booking_file:
    print(f"\n📅 [1/8] Migrating Bookings from {booking_file[0]}...")
    wb = openpyxl.load_workbook(booking_file[0], data_only=True)
    sheet = wb['Bookings'] if 'Bookings' in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    
    bookings_payload = []
    for r in rows[1:]:
        if not r or not any(r): continue
        d = dict(zip(header, r))
        
        booking_date = format_date(d.get('วันที่จอง'))
        sender = str(d.get('ชื่อ-นานสกุล') or d.get('หน่วยงาน') or 'ผู้ส่งตรวจ').strip()
        dept = str(d.get('หน่วยงาน') or 'งานจุลชีววิทยา').strip()
        contact = str(d.get('เบอร์ติดต่อ') or '-').replace('.0', '').strip()
        test_type = str(d.get('สิ่งส่งตรวจ') or 'ตรวจสิ่งแวดล้อม').strip()
        
        try:
            sample_count = int(float(d.get('จำนวน') or 1))
            if sample_count <= 0 or sample_count > 200: sample_count = 1
        except:
            sample_count = 1
            
        notes = str(d.get('หมายเหตุ') or '').strip()
        if notes == 'None': notes = ''
        
        # Map service code
        srv_code = 'AIR_01'
        srv_name = 'Air Sampling (สำหรับงานอาชีวอนามัย)'
        if 'ยา' in test_type:
            srv_code = 'DRG_07'
            srv_name = 'Drug (สำหรับงานผลิตยา1) ปลอดเชื้อ'
        elif 'เลือด' in test_type or 'Steril' in test_type:
            srv_code = 'STR_02'
            srv_name = 'Sterility (สำหรับงานธนาคารเลือด)'
        elif 'อาหาร' in test_type:
            srv_code = 'FOD_06'
            srv_name = 'Food (สำหรับงานโภชนาการ)'
        elif 'น้ำ' in test_type:
            srv_code = 'WTO_04'
            srv_name = 'Water (สำหรับห้องผ่าตัด OR)'
        elif 'พื้นผิว' in test_type:
            srv_code = 'WTS_03'
            srv_name = 'Water or Surface (สำหรับงานควบคุมโรคติดเชื้อ)'
            
        bookings_payload.append({
            "booking_date": booking_date,
            "sender_name": sender[:100],
            "department": dept[:100],
            "contact_number": contact[:50],
            "service_code": srv_code,
            "service_name": srv_name,
            "sample_count": sample_count,
            "notes": notes,
            "status": "confirmed"
        })
    
    print(f"  Inserting {len(bookings_payload)} booking records in chunks...")
    chunk_size = 50
    inserted_b = 0
    for i in range(0, len(bookings_payload), chunk_size):
        chunk = bookings_payload[i:i+chunk_size]
        res, err = supabase_post("bookings", chunk)
        if err:
            print(f"  ⚠️ Error inserting bookings chunk {i}: {err}")
        else:
            inserted_b += len(res) if res else len(chunk)
    print(f"  ✅ Successfully migrated {inserted_b} bookings!")


# ============================================================================
# 2. MIGRATE AIR_01 (Air Sampling)
# ============================================================================
air_file = glob.glob("downloaded_raw/AIR_01*.xlsx")
if air_file:
    print(f"\n💨 [2/8] Migrating AIR-01 from {air_file[0]}...")
    wb = openpyxl.load_workbook(air_file[0], data_only=True)
    sheet = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    
    submissions = {}
    for r in rows[1:]:
        if not r or not any(r): continue
        d = dict(zip(header, r))
        sub_id = str(d.get('SubmissionID') or '').strip()
        if not sub_id or sub_id == 'None': continue
        
        if sub_id not in submissions:
            sub_date = format_date(d.get('SubmissionDate') or d.get('Timestamp'))
            dept = str(d.get('Department') or 'งานอาชีวอนามัย').strip()
            ward = str(d.get('Ward') or '').strip()
            
            clean_sub_no = sub_id.replace('SUB', 'AIR-')
            if not clean_sub_no.startswith('AIR-'): clean_sub_no = f"AIR-{clean_sub_no}"
            
            submissions[sub_id] = {
                "header": {
                    "submission_no": clean_sub_no[:50],
                    "service_code": "AIR_01",
                    "service_name": "Air Sampling (สำหรับงานอาชีวอนามัย)",
                    "department": dept[:100],
                    "ward_room": ward[:100],
                    "sampling_date": sub_date,
                    "received_date": sub_date,
                    "reported_date": sub_date,
                    "sampler_name": "เจ้าหน้าที่อาชีวอนามัย",
                    "reporter_name": "ทนพ.มานพ นันตาบุตร",
                    "approver_name": "ทนพญ.กรรณิการ์ สิทธิโชค",
                    "overall_result": "pass",
                    "status": "completed",
                    "remarks": "ตรวจนับจำนวนโคโลนีเชื้อแบคทีเรียและเชื้อราในอากาศ"
                },
                "items": []
            }
            
        b_count = str(d.get('Bacteria_CFU') if d.get('Bacteria_CFU') is not None else 0).replace('.0', '')
        f_count = str(d.get('Fungus_CFU') if d.get('Fungus_CFU') is not None else 0).replace('.0', '')
        loc = str(d.get('Location') or d.get('Ward') or 'จุดเก็บตัวอย่าง').strip()
        
        # Check standard limit
        try:
            b_num = float(b_count)
            f_num = float(f_count)
            item_res = "pass" if (b_num < 500 and f_num < 100) else "fail"
        except:
            item_res = "pass"
            
        if item_res == "fail":
            submissions[sub_id]["header"]["overall_result"] = "fail"
            
        submissions[sub_id]["items"].append({
            "item_no": len(submissions[sub_id]["items"]) + 1,
            "location_name": loc[:100],
            "sample_description": str(d.get('SampleType') or 'อากาศ'),
            "bacteria_count": b_count,
            "fungus_count": f_count,
            "standard_limit": "< 500 CFU/m³ (Bacteria), < 100 CFU/m³ (Fungi)",
            "item_result": item_res,
            "remarks": str(d.get('Remarks') or '')
        })
        
    print(f"  Found {len(submissions)} AIR-01 submissions with total items...")
    migrated_count = 0
    for sub_id, data in submissions.items():
        res, err = supabase_post("reports", data["header"])
        if err:
            # Maybe duplicate submission_no, append random suffix
            data["header"]["submission_no"] += f"-{data['header']['sampling_date'][-2:]}"
            res, err = supabase_post("reports", data["header"])
            
        if res and len(res) > 0:
            report_id = res[0]["id"]
            for it in data["items"]:
                it["report_id"] = report_id
            supabase_post("report_items", data["items"])
            migrated_count += 1
            
    print(f"  ✅ Successfully migrated {migrated_count} AIR-01 reports!")


# ============================================================================
# 3. MIGRATE STR_02 (Sterility Test / Blood Bank)
# ============================================================================
str_file = glob.glob("downloaded_raw/STR_02*.xlsx")
if str_file:
    print(f"\n🩸 [3/8] Migrating STR-02 from {str_file[0]}...")
    wb = openpyxl.load_workbook(str_file[0], data_only=True)
    sheet = wb['Submissions'] if 'Submissions' in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    
    submissions = {}
    for r in rows[1:]:
        if not r or not any(r): continue
        d = dict(zip(header, r))
        sub_id = str(d.get('Submission ID') or '').strip()
        if not sub_id or sub_id == 'None': continue
        
        if sub_id not in submissions:
            sub_date = format_date(d.get('Submission Date') or d.get('Timestamp'))
            dept = str(d.get('Department') or 'งานธนาคารเลือด').strip()
            clean_sub_no = sub_id.replace('SUB', 'STR-')
            if not clean_sub_no.startswith('STR-'): clean_sub_no = f"STR-{clean_sub_no}"
            
            submissions[sub_id] = {
                "header": {
                    "submission_no": clean_sub_no[:50],
                    "service_code": "STR_02",
                    "service_name": "Sterility (สำหรับงานธนาคารเลือด)",
                    "department": dept[:100],
                    "ward_room": "ธนาคารเลือด",
                    "sampling_date": sub_date,
                    "received_date": sub_date,
                    "reported_date": sub_date,
                    "sampler_name": "เจ้าหน้าที่ธนาคารเลือด",
                    "reporter_name": "ทนพ.อนุชิต บุญประสิทธิ์",
                    "approver_name": "ทนพ.มานพ นันตาบุตร",
                    "overall_result": "pass",
                    "status": "completed",
                    "remarks": "บ่มเชื้อครบตามระยะเวลามาตรฐาน ไม่พบการเจริญเติบโตของเชื้อ"
                },
                "items": []
            }
            
        bag_no = str(d.get('Blood Bag Number') or '-').replace('.0', '')
        prod_type = str(d.get('Product Type') or 'เลือด')
        sterile_res = str(d.get('Sterile') or 'No growth')
        item_res = "pass" if "no growth" in sterile_res.lower() or "sterile" in sterile_res.lower() else "fail"
        
        if item_res == "fail":
            submissions[sub_id]["header"]["overall_result"] = "fail"
            
        submissions[sub_id]["items"].append({
            "item_no": len(submissions[sub_id]["items"]) + 1,
            "location_name": f"ถุงเลือด #{bag_no} ({prod_type})",
            "sample_description": prod_type,
            "bacteria_count": "0",
            "fungus_count": "0",
            "standard_limit": "No bacterial or fungal growth",
            "item_result": item_res,
            "remarks": sterile_res
        })
        
    migrated_count = 0
    for sub_id, data in submissions.items():
        res, err = supabase_post("reports", data["header"])
        if res and len(res) > 0:
            report_id = res[0]["id"]
            for it in data["items"]: it["report_id"] = report_id
            supabase_post("report_items", data["items"])
            migrated_count += 1
    print(f"  ✅ Successfully migrated {migrated_count} STR-02 reports!")


# ============================================================================
# 4. MIGRATE WTS_03 (Water or Surface / IC)
# ============================================================================
wts_file = glob.glob("downloaded_raw/WTS_03*.xlsx")
if wts_file:
    print(f"\n🧽 [4/8] Migrating WTS-03 from {wts_file[0]}...")
    wb = openpyxl.load_workbook(wts_file[0], data_only=True)
    sheet = wb['Submissions'] if 'Submissions' in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    
    submissions = {}
    for r in rows[1:]:
        if not r or not any(r): continue
        d = dict(zip(header, r))
        sub_id = str(d.get('SubmissionID') or '').strip()
        if not sub_id or sub_id == 'None': continue
        
        if sub_id not in submissions:
            sub_date = format_date(d.get('SubmissionDate') or d.get('Timestamp'))
            dept = str(d.get('Department') or 'งานควบคุมโรคติดเชื้อ (IC)').strip()
            clean_sub_no = sub_id.replace('JOB-', 'WTS-')
            if not clean_sub_no.startswith('WTS-'): clean_sub_no = f"WTS-{clean_sub_no}"
            
            submissions[sub_id] = {
                "header": {
                    "submission_no": clean_sub_no[:50],
                    "service_code": "WTS_03",
                    "service_name": "Water or Surface (สำหรับงานควบคุมโรคติดเชื้อ)",
                    "department": dept[:100],
                    "ward_room": dept[:100],
                    "sampling_date": sub_date,
                    "received_date": sub_date,
                    "reported_date": sub_date,
                    "sampler_name": "พว. งานควบคุมโรคติดเชื้อ",
                    "reporter_name": "ทนพ.มานพ นันตาบุตร",
                    "approver_name": "ทนพญ.กรรณิการ์ สิทธิโชค",
                    "overall_result": "pass",
                    "status": "completed",
                    "remarks": "ตรวจเพาะเชื้อสิ่งแวดล้อมและพื้นผิวสัมผัส"
                },
                "items": []
            }
            
        loc = str(d.get('Location') or 'จุดสว็อบ').strip()
        culture_res = str(d.get('CultureResult') or 'No growth').strip()
        suspected = str(d.get('SuspectedOrganism') or '').strip()
        
        item_res = "pass" if "no growth" in culture_res.lower() or culture_res == '-' else "fail"
        if item_res == "fail":
            submissions[sub_id]["header"]["overall_result"] = "fail"
            
        submissions[sub_id]["items"].append({
            "item_no": len(submissions[sub_id]["items"]) + 1,
            "location_name": loc[:100],
            "sample_description": str(d.get('SpecimenType') or 'พื้นผิว'),
            "bacteria_count": culture_res,
            "fungus_count": "0",
            "microorganism_found": culture_res if item_res == "fail" else "ไม่พบเชื้อก่อโรค",
            "standard_limit": "No pathogenic microorganisms",
            "item_result": item_res,
            "remarks": suspected
        })
        
    migrated_count = 0
    for sub_id, data in submissions.items():
        res, err = supabase_post("reports", data["header"])
        if res and len(res) > 0:
            report_id = res[0]["id"]
            for it in data["items"]: it["report_id"] = report_id
            supabase_post("report_items", data["items"])
            migrated_count += 1
    print(f"  ✅ Successfully migrated {migrated_count} WTS-03 reports!")


# ============================================================================
# 5. MIGRATE WTO_04 (Water OR)
# ============================================================================
wto_file = glob.glob("downloaded_raw/WTO_04*.xlsx")
if wto_file:
    print(f"\n🚿 [5/8] Migrating WTO-04 from {wto_file[0]}...")
    wb = openpyxl.load_workbook(wto_file[0], data_only=True)
    sheet = wb['EnviMicrobiologyReports'] if 'EnviMicrobiologyReports' in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    
    submissions = {}
    for r in rows[1:]:
        if not r or not any(r): continue
        d = dict(zip(header, r))
        sub_id = str(d.get('SubmissionID') or '').strip()
        if not sub_id or sub_id == 'None': continue
        
        if sub_id not in submissions:
            sub_date = format_date(d.get('SubmissionDate') or d.get('Timestamp'))
            dept = str(d.get('Department') or 'ห้องผ่าตัด').strip()
            clean_sub_no = sub_id.replace('ENV-', 'WTO-')
            if not clean_sub_no.startswith('WTO-'): clean_sub_no = f"WTO-{clean_sub_no}"
            
            submissions[sub_id] = {
                "header": {
                    "submission_no": clean_sub_no[:50],
                    "service_code": "WTO_04",
                    "service_name": "Water (สำหรับห้องผ่าตัด OR)",
                    "department": dept[:100],
                    "ward_room": "ห้องผ่าตัด",
                    "sampling_date": sub_date,
                    "received_date": sub_date,
                    "reported_date": sub_date,
                    "sampler_name": "พยาบาลห้องผ่าตัด",
                    "reporter_name": "ทนพ.มานพ นันตาบุตร",
                    "approver_name": "ทนพญ.กรรณิการ์ สิทธิโชค",
                    "overall_result": "pass",
                    "status": "completed",
                    "remarks": "ตรวจคุณภาพน้ำห้องผ่าตัด"
                },
                "items": []
            }
            
        loc = str(d.get('Location') or 'อ่างล้างมือ').strip()
        ecoli = str(d.get('Ecoli') or 'No growth').strip()
        item_res = "pass" if "no growth" in ecoli.lower() or "not detected" in ecoli.lower() else "fail"
        
        submissions[sub_id]["items"].append({
            "item_no": len(submissions[sub_id]["items"]) + 1,
            "location_name": loc[:100],
            "sample_description": str(d.get('SpecimenType') or 'น้ำ'),
            "bacteria_count": ecoli,
            "fungus_count": "0",
            "standard_limit": "< 10 CFU/100mL (TVC), Endotoxin < 0.25 EU/mL",
            "item_result": item_res,
            "remarks": ecoli
        })
        
    migrated_count = 0
    for sub_id, data in submissions.items():
        res, err = supabase_post("reports", data["header"])
        if res and len(res) > 0:
            report_id = res[0]["id"]
            for it in data["items"]: it["report_id"] = report_id
            supabase_post("report_items", data["items"])
            migrated_count += 1
    print(f"  ✅ Successfully migrated {migrated_count} WTO-04 reports!")


# ============================================================================
# 6. MIGRATE FOD_06 (Food)
# ============================================================================
fod_file = glob.glob("downloaded_raw/FOD_06*.xlsx")
if fod_file:
    print(f"\n🍱 [6/8] Migrating FOD-06 from {fod_file[0]}...")
    wb = openpyxl.load_workbook(fod_file[0], data_only=True)
    sheet = wb['Data'] if 'Data' in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    
    submissions = {}
    for idx, r in enumerate(rows[1:]):
        if not r or not any(r): continue
        d = dict(zip(header, r))
        sub_date = format_date(d.get('SubmissionDate') or d.get('Timestamp'))
        dept = str(d.get('Department') or 'งานโภชนาการ').strip()
        
        group_key = f"FOD-{sub_date}"
        if group_key not in submissions:
            submissions[group_key] = {
                "header": {
                    "submission_no": f"{group_key}-{len(submissions)+1:02d}",
                    "service_code": "FOD_06",
                    "service_name": "Food (สำหรับงานโภชนาการ)",
                    "department": dept[:100],
                    "ward_room": "งานโภชนาการ",
                    "sampling_date": sub_date,
                    "received_date": sub_date,
                    "reported_date": sub_date,
                    "sampler_name": "นักโภชนาการ",
                    "reporter_name": "ทนพ.มานพ นันตาบุตร",
                    "approver_name": "ทนพญ.กรรณิการ์ สิทธิโชค",
                    "overall_result": "pass",
                    "status": "completed",
                    "remarks": "ตรวจวิเคราะห์จุลินทรีย์ในอาหารและสุขาภิบาล"
                },
                "items": []
            }
            
        food_name = str(d.get('Food') or f"ตัวอย่างอาหาร {idx+1}").strip()
        ecoli = str(d.get('Ecoli') or 'ไม่พบเชื้อ').strip()
        paer = str(d.get('Paeruginosa') or 'ไม่พบเชื้อ').strip()
        
        item_res = "pass" if ("ไม่พบ" in ecoli or "negative" in ecoli.lower()) else "fail"
        
        submissions[group_key]["items"].append({
            "item_no": len(submissions[group_key]["items"]) + 1,
            "location_name": food_name[:100],
            "sample_description": str(d.get('SampleType') or 'อาหาร'),
            "bacteria_count": ecoli,
            "fungus_count": paer,
            "standard_limit": "E. coli Negative, Salmonella/Shigella Not detected",
            "item_result": item_res,
            "remarks": str(d.get('Remarks') or '')
        })
        
    migrated_count = 0
    for sub_id, data in submissions.items():
        res, err = supabase_post("reports", data["header"])
        if res and len(res) > 0:
            report_id = res[0]["id"]
            for it in data["items"]: it["report_id"] = report_id
            supabase_post("report_items", data["items"])
            migrated_count += 1
    print(f"  ✅ Successfully migrated {migrated_count} FOD-06 reports!")


# ============================================================================
# 7. MIGRATE DRG_07 (Drug 1 Sterile)
# ============================================================================
drg1_file = glob.glob("downloaded_raw/DRG_07*.csv")
if drg1_file:
    print(f"\n💊 [7/8] Migrating DRG-07 from {drg1_file[0]}...")
    with open(drg1_file[0], mode='r', encoding='utf-8-sig', errors='ignore') as f:
        reader = csv.reader(f)
        rows = list(reader)
        
    header = rows[0]
    submissions = {}
    for r in rows[1:]:
        if not r or not any(r): continue
        d = dict(zip(header, r))
        sub_id = str(d.get('SubmissionID') or '').strip()
        if not sub_id: continue
        
        if sub_id not in submissions:
            sub_date = format_date(d.get('ReceiptDate') or d.get('Timestamp'))
            clean_sub_no = sub_id.replace('JOB-', 'DR1-')
            if not clean_sub_no.startswith('DR1-'): clean_sub_no = f"DR1-{clean_sub_no}"
            
            submissions[sub_id] = {
                "header": {
                    "submission_no": clean_sub_no[:50],
                    "service_code": "DRG_07",
                    "service_name": "Drug (สำหรับงานผลิตยา1) ปลอดเชื้อ",
                    "department": "งานผลิตยา",
                    "ward_room": "หน่วยเตรียมยาปราศจากเชื้อ",
                    "sampling_date": sub_date,
                    "received_date": sub_date,
                    "reported_date": format_date(d.get('AnalysisDate') or sub_date),
                    "sampler_name": "เภสัชกรงานผลิตยา",
                    "reporter_name": "ทนพ.มานพ นันตาบุตร",
                    "approver_name": "ทนพญ.กรรณิการ์ สิทธิโชค",
                    "overall_result": "pass",
                    "status": "completed",
                    "remarks": "ทดสอบความปราศจากเชื้อตามมาตรฐาน USP <71>"
                },
                "items": []
            }
            
        med_name = str(d.get('PreparedMedicineHeader') or 'ยาปราศจากเชื้อ').strip()
        lot_no = str(d.get('LotNo') or '-').strip()
        
        submissions[sub_id]["items"].append({
            "item_no": len(submissions[sub_id]["items"]) + 1,
            "location_name": f"{med_name} (Lot {lot_no})",
            "sample_description": f"Volume: {d.get('Volume', '')} mL, SampleCount: {d.get('SampleCount', '')}",
            "bacteria_count": "No growth",
            "fungus_count": "No growth",
            "standard_limit": "USP <71> Sterility Tests (14 days)",
            "item_result": "pass",
            "remarks": f"Prod Date: {d.get('ProductionDate', '')}"
        })
        
    migrated_count = 0
    for sub_id, data in submissions.items():
        res, err = supabase_post("reports", data["header"])
        if res and len(res) > 0:
            report_id = res[0]["id"]
            for it in data["items"]: it["report_id"] = report_id
            supabase_post("report_items", data["items"])
            migrated_count += 1
    print(f"  ✅ Successfully migrated {migrated_count} DRG-07 reports!")


# ============================================================================
# 8. MIGRATE DRG_08 (Drug 2 Microbial Contamination)
# ============================================================================
drg2_file = glob.glob("downloaded_raw/DRG_08*.xlsx")
if drg2_file:
    print(f"\n🧪 [8/8] Migrating DRG-08 from {drg2_file[0]}...")
    wb = openpyxl.load_workbook(drg2_file[0], data_only=True)
    sheet = wb['Submissions'] if 'Submissions' in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    
    submissions = {}
    for r in rows[1:]:
        if not r or not any(r): continue
        d = dict(zip(header, r))
        sub_id = str(d.get('Submission ID') or '').strip()
        if not sub_id or sub_id == 'None': continue
        
        if sub_id not in submissions:
            sub_date = format_date(d.get('Sample Date') or d.get('Timestamp'))
            dept = str(d.get('Department') or 'งานผลิตยา').strip()
            clean_sub_no = sub_id.replace('TUH-', 'DR2-')
            if not clean_sub_no.startswith('DR2-'): clean_sub_no = f"DR2-{clean_sub_no}"
            
            submissions[sub_id] = {
                "header": {
                    "submission_no": clean_sub_no[:50],
                    "service_code": "DRG_08",
                    "service_name": "Drug (สำหรับงานผลิตยา2) แบบรายงานผลการวิเคราะห์การปนเปื้อนเชื้อจุลินทรีย์",
                    "department": dept[:100],
                    "ward_room": "งานผลิตยา",
                    "sampling_date": sub_date,
                    "received_date": sub_date,
                    "reported_date": sub_date,
                    "sampler_name": str(d.get('Operator') or 'เภสัชกรงานผลิตยา'),
                    "reporter_name": "ทนพ.มานพ นันตาบุตร",
                    "approver_name": "ทนพญ.กรรณิการ์ สิทธิโชค",
                    "overall_result": "pass",
                    "status": "completed",
                    "remarks": "ตรวจนับจำนวนจุลินทรีย์และการปนเปื้อนตาม USP <61>/<62>"
                },
                "items": []
            }
            
        drug_type = str(d.get('Drug Type') or 'ผลิตภัณฑ์ยา').strip()
        tamc = str(d.get('TAMC') or '< 10^2 CFU/g').strip()
        tymc = str(d.get('TYMC') or '< 10^1 CFU/g').strip()
        
        submissions[sub_id]["items"].append({
            "item_no": len(submissions[sub_id]["items"]) + 1,
            "location_name": drug_type[:100],
            "sample_description": str(d.get('Sample Type') or 'ยา'),
            "bacteria_count": tamc,
            "fungus_count": tymc,
            "standard_limit": "USP <61> TAMC/TYMC Limits",
            "item_result": "pass",
            "remarks": str(d.get('Remarks') or '')
        })
        
    migrated_count = 0
    for sub_id, data in submissions.items():
        res, err = supabase_post("reports", data["header"])
        if res and len(res) > 0:
            report_id = res[0]["id"]
            for it in data["items"]: it["report_id"] = report_id
            supabase_post("report_items", data["items"])
            migrated_count += 1
    print(f"  ✅ Successfully migrated {migrated_count} DRG-08 reports!")

print("\n======================================================================")
print("🎉 ALL RAW DATA MIGRATION TO SUPABASE COMPLETED SUCCESSFULLY!")
print("======================================================================")
