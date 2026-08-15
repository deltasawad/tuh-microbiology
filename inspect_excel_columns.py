import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
import csv
import os
import glob

print("=== INSPECTING DOWNLOADED RAW FILES ===")

files = glob.glob("downloaded_raw/*.*")
for fpath in sorted(files):
    fname = os.path.basename(fpath)
    print(f"\n-----------------------------------------------------------")
    print(f"📁 File: {fname}")
    print(f"-----------------------------------------------------------")
    
    if fpath.endswith('.xlsx'):
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
            print(f"Sheets: {wb.sheetnames}")
            for sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                rows = list(sheet.iter_rows(values_only=True))
                print(f"  Sheet '{sheetname}': Total rows = {len(rows)}")
                if len(rows) > 0:
                    print(f"    Header (Row 1): {rows[0][:10]}")
                if len(rows) > 1:
                    print(f"    Sample (Row 2): {rows[1][:10]}")
        except Exception as e:
            print(f"Error reading xlsx: {e}")
            
    elif fpath.endswith('.csv'):
        try:
            with open(fpath, "r", encoding="utf-8-sig", errors="ignore") as f:
                reader = csv.reader(f)
                rows = list(reader)
                print(f"  CSV Total rows = {len(rows)}")
                if len(rows) > 0:
                    print(f"    Header: {rows[0][:10]}")
                if len(rows) > 1:
                    print(f"    Sample: {rows[1][:10]}")
        except Exception as e:
            print(f"Error reading csv: {e}")
