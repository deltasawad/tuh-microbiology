import os
import sys
import openpyxl
from datetime import datetime, date
from sqlalchemy.orm import Session

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from app.core.database import SessionLocal, Base, engine
from app.models.master import Service, Department, Ward, Staff
from app.models.submission import Submission
from app.models.sample import Sample, SampleResult
from app.models.status_transition import StatusTransition
from app.models.audit import AuditLog
from app.models.booking import Booking
from app.models.user import User

def parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    elif isinstance(val, date):
        return val
    elif isinstance(val, str):
        try:
            return datetime.strptime(val.strip(), "%Y-%m-%d").date()
        except Exception:
            try:
                return datetime.strptime(val.strip(), "%d/%m/%Y").date()
            except Exception:
                pass
    return date.today()

def find_excel_file(keyword: str) -> str | None:
    search_dirs = [
        os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..")),
        os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")),
        os.path.abspath("."),
        os.path.abspath("..")
    ]
    for d in search_dirs:
        if os.path.exists(d):
            for root, _, files in os.walk(d):
                for f in files:
                    if f.endswith(".xlsx") and (keyword.lower() in f.lower() or keyword in f):
                        return os.path.join(root, f)
    return None

def migrate_real_data():
    print("==========================================================")
    print("  MIGRATING REAL DATA FROM EXCEL TO SUPABASE DATABASE")
    print("==========================================================")
    
    db: Session = SessionLocal()

    try:
        air_service = db.query(Service).filter(Service.code == "AIR-01").first()
        if not air_service:
            air_service = Service(code="AIR-01", name_th="Air Sampling (ตรวจคุณภาพอากาศ)", department_owner="งานอาชีวอนามัยและความปลอดภัย", tat_target_hours=120)
            db.add(air_service)
            db.flush()

        default_dept = db.query(Department).filter(Department.name_th.like("%อาชีวอนามัย%")).first()
        if not default_dept:
            default_dept = db.query(Department).first()

        admin_user = db.query(User).filter(User.role == "admin").first()
        staff_manop = db.query(Staff).filter(Staff.first_name == "มานพ").first()
        staff_narisara = db.query(Staff).filter(Staff.first_name == "นริศรา").first()

        # ---------------------------------------------------------
        # 1. Migrate Air Sampling Excel Data
        # ---------------------------------------------------------
        air_file = find_excel_file("Air Sampling")
        if not air_file:
            air_file = find_excel_file("air")
        
        if air_file and os.path.exists(air_file):
            print(f"\n[1/2] Loading Air Sampling real data from: {os.path.basename(air_file)}")
            wb = openpyxl.load_workbook(air_file, data_only=True)
            sheet = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

            submissions_dict = {}
            for r in range(2, sheet.max_row + 1):
                sub_id_val = sheet.cell(r, 2).value
                if not sub_id_val:
                    continue
                sub_id = str(sub_id_val).strip()
                if sub_id not in submissions_dict:
                    submissions_dict[sub_id] = []

                submissions_dict[sub_id].append({
                    "timestamp": sheet.cell(r, 1).value,
                    "sub_date": sheet.cell(r, 3).value,
                    "dept_name": sheet.cell(r, 4).value,
                    "sample_type": sheet.cell(r, 5).value or "อากาศ",
                    "sample_no": sheet.cell(r, 6).value or len(submissions_dict[sub_id]) + 1,
                    "ward_name": sheet.cell(r, 7).value,
                    "location": sheet.cell(r, 8).value or "",
                    "bac_cfu": sheet.cell(r, 9).value,
                    "fun_cfu": sheet.cell(r, 10).value,
                    "status": sheet.cell(r, 11).value or "ตรวจแล้ว",
                    "remarks": sheet.cell(r, 12).value or "",
                    "email": sheet.cell(r, 14).value or ""
                })

            print(f"  Found {len(submissions_dict)} distinct Air Sampling Submissions.")

            sub_count = 0
            sample_count = 0

            for sub_key, rows in submissions_dict.items():
                first_row = rows[0]
                dept_str = str(first_row["dept_name"]).strip() if first_row["dept_name"] else "งานอาชีวอนามัยและความปลอดภัย"
                
                dept = db.query(Department).filter(Department.name_th == dept_str).first()
                if not dept:
                    dept = Department(name_th=dept_str)
                    db.add(dept)
                    db.flush()

                sub_date = parse_date(first_row["sub_date"])
                raw_status = str(first_row["status"]).strip()
                final_status = "REPORTED" if ("ตรวจแล้ว" in raw_status or "ออกผล" in raw_status) else "SUBMITTED"

                existing_sub = db.query(Submission).filter(Submission.submission_no == sub_key).first()
                if not existing_sub:
                    submission = Submission(
                        submission_no=sub_key,
                        service_id=air_service.id,
                        department_id=dept.id,
                        submission_date=sub_date,
                        sender_name="เจ้าหน้าที่ผู้ส่งตรวจ",
                        sender_email=first_row["email"] or None,
                        sample_type=first_row["sample_type"] or "อากาศ",
                        sample_count=len(rows),
                        status=final_status,
                        reporter_id=staff_manop.id if (final_status == "REPORTED" and staff_manop) else None,
                        reviewer_id=staff_narisara.id if (final_status == "REPORTED" and staff_narisara) else None,
                        reported_at=datetime.combine(sub_date, datetime.min.time()) if final_status == "REPORTED" else None,
                        reviewed_at=datetime.combine(sub_date, datetime.min.time()) if final_status == "REPORTED" else None
                    )
                    db.add(submission)
                    db.flush()
                    sub_count += 1

                    db.add(StatusTransition(
                        submission_id=submission.id,
                        from_status="DRAFT",
                        to_status="SUBMITTED",
                        actor_user_id=admin_user.id if admin_user else None,
                        reason="นำเข้าข้อมูลจริงจากระบบเดิม"
                    ))
                    if final_status == "REPORTED":
                        db.add(StatusTransition(
                            submission_id=submission.id,
                            from_status="SUBMITTED",
                            to_status="REPORTED",
                            actor_user_id=admin_user.id if admin_user else None,
                            reason="นำเข้าผลตรวจเสร็จสิ้นจากระบบเดิม"
                        ))

                    for idx, s_row in enumerate(rows, 1):
                        ward_str = str(s_row["ward_name"]).strip() if s_row["ward_name"] else "-"
                        ward = db.query(Ward).filter(Ward.name_th == ward_str).first()
                        if not ward and ward_str != "-":
                            ward = Ward(name_th=ward_str, department_id=dept.id)
                            db.add(ward)
                            db.flush()

                        sample = Sample(
                            submission_id=submission.id,
                            sample_no=idx,
                            ward_id=ward.id if ward else None,
                            label=str(s_row["location"]).strip() if s_row["location"] else None
                        )
                        db.add(sample)
                        db.flush()
                        sample_count += 1

                        # Bacteria Result
                        if s_row["bac_cfu"] is not None:
                            try:
                                bac_num = float(s_row["bac_cfu"])
                                flag = "ABNORMAL" if bac_num > 50 else "NORMAL"
                            except Exception:
                                bac_num = None
                                flag = "NORMAL"

                            db.add(SampleResult(
                                sample_id=sample.id,
                                analyte_code="bacteria_colonies",
                                result_value=str(s_row["bac_cfu"]),
                                numeric_value=bac_num,
                                result_flag=flag,
                                remarks=str(s_row["remarks"]).strip() if s_row["remarks"] else None
                            ))

                        # Fungus Result
                        if s_row["fun_cfu"] is not None:
                            try:
                                fun_num = float(s_row["fun_cfu"])
                                flag = "ABNORMAL" if fun_num > 0 else "NORMAL"
                            except Exception:
                                fun_num = None
                                flag = "NORMAL"

                            db.add(SampleResult(
                                sample_id=sample.id,
                                analyte_code="fungus_colonies",
                                result_value=str(s_row["fun_cfu"]),
                                numeric_value=fun_num,
                                result_flag=flag
                            ))

            print(f"  --> Successfully imported {sub_count} Submissions with {sample_count} Sample points!")

        # ---------------------------------------------------------
        # 2. Migrate Booking Calendar Excel Data
        # ---------------------------------------------------------
        book_file = find_excel_file("จองคิว")
        if not book_file:
            book_file = find_excel_file("book")
        
        if book_file and os.path.exists(book_file):
            print(f"\n[2/2] Loading Booking calendar real data from: {os.path.basename(book_file)}")
            wb_b = openpyxl.load_workbook(book_file, data_only=True)
            sheet_b = wb_b["Bookings"] if "Bookings" in wb_b.sheetnames else wb_b.active

            book_count = 0
            for r in range(2, sheet_b.max_row + 1):
                b_date_raw = sheet_b.cell(r, 1).value
                if not b_date_raw:
                    continue

                b_date = parse_date(b_date_raw)
                name = str(sheet_b.cell(r, 2).value or "ผู้ส่งตรวจ").strip()
                dept_name = str(sheet_b.cell(r, 3).value or "").strip()
                phone = str(sheet_b.cell(r, 4).value or "-").replace(".0", "").strip()
                test_type = str(sheet_b.cell(r, 5).value or "อากาศ").strip()
                
                try:
                    count_val = int(float(sheet_b.cell(r, 6).value or 1))
                except Exception:
                    count_val = 1

                remarks = str(sheet_b.cell(r, 9).value or "").strip()
                dept = db.query(Department).filter(Department.name_th == dept_name).first()

                existing_book = db.query(Booking).filter(
                    Booking.booking_date == b_date,
                    Booking.full_name == name,
                    Booking.test_type == test_type
                ).first()

                if not existing_book:
                    booking = Booking(
                        booking_date=b_date,
                        test_type=test_type,
                        department_id=dept.id if dept else None,
                        department_name=dept_name or (dept.name_th if dept else None),
                        full_name=name,
                        contact_number=phone,
                        sample_count=count_val,
                        remarks=remarks or None,
                        status="CONFIRMED"
                    )
                    db.add(booking)
                    book_count += 1

            print(f"  --> Successfully imported {book_count} Real Booking records into Calendar!")

        db.commit()
        print("\n==========================================================")
        print("  ALL REAL HOSPITAL DATA SUCCESSFULLY COMMITTED TO SUPABASE!")
        print("==========================================================")

    except Exception as e:
        db.rollback()
        print(f"Error migrating real data: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()

if __name__ == "__main__":
    migrate_real_data()
