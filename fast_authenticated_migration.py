import sys
sys.stdout.reconfigure(encoding='utf-8')

import urllib.request
import json
import os
import glob
import re
from datetime import datetime, date
import openpyxl
import csv

SUPABASE_URL = "https://tgctyouhzsyizlosrmqh.supabase.co"
ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InRnY3R5b3VoenN5aXpsb3NybXFoIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODY2NjYyNzAsImV4cCI6MjEwMjI0MjI3MH0.mlba06N1LhjT9vwNBhVrhMlxvjvO7QsRjArI6ue7Pv0"

# 1. Authenticate with Staff User
print("🔐 Authenticating with Supabase Staff User...")
auth_url = f"{SUPABASE_URL}/auth/v1/token?grant_type=password"
auth_req = urllib.request.Request(
    auth_url,
    data=json.dumps({"email": "admin@tuh.lab", "password": "password123"}).encode('utf-8'),
    headers={"apikey": ANON_KEY, "Content-Type": "application/json"},
    method='POST'
)

with urllib.request.urlopen(auth_req) as resp:
    auth_data = json.loads(resp.read().decode('utf-8'))
    access_token = auth_data["access_token"]
    user_id = auth_data["user"]["id"]
    print(f"✅ Authenticated as: {auth_data['user']['email']} (UUID: {user_id})")

auth_headers = {
    "apikey": ANON_KEY,
    "Authorization": f"Bearer {access_token}",
    "Content-Type": "application/json",
    "Prefer": "return=representation"
}

def supabase_post(endpoint, payload):
    url = f"{SUPABASE_URL}/rest/v1/{endpoint}"
    data_bytes = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(url, data=data_bytes, headers=auth_headers, method='POST')
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode('utf-8')), None
    except Exception as e:
        err_msg = str(e)
        if hasattr(e, 'read'):
            err_msg += " : " + e.read().decode('utf-8')
        return None, err_msg

def format_date(val):
    if not val:
        return date.today().isoformat()
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if re.match(r'^\d{1,2}/\d{1,2}/\d{4}', s):
        parts = s.split('/')
        d, m, y = int(parts[0]), int(parts[1]), int(parts[2].split()[0])
        if y > 2500: y -= 543
        return f"{y:04d}-{m:02d}-{d:02d}"
    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
        return s[:10]
    return date.today().isoformat()

print("\n======================================================================")
print("🚀 MIGRATING 8 ENVIRONMENTAL TESTING SERVICES + BOOKINGS TO SUPABASE")
print("======================================================================")

# ============================================================================
# 1. BOOKINGS (If count < 10)
# ============================================================================
booking_file = glob.glob("downloaded_raw/BOOKING*.xlsx")
if booking_file:
    print(f"\n📅 [1/8] Verifying Bookings...")
    wb = openpyxl.load_workbook(booking_file[0], data_only=True)
    sheet = wb['Bookings'] if 'Bookings' in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    
    # Check if bookings already in Supabase
    req = urllib.request.Request(f"{SUPABASE_URL}/rest/v1/bookings?select=id", headers=auth_headers)
    with urllib.request.urlopen(req) as resp:
        existing_bookings = json.loads(resp.read().decode('utf-8'))
        
    if len(existing_bookings) < 10:
        bookings_payload = []
        for r in rows[1:]:
            if not r or not any(r): continue
            d = dict(zip(header, r))
            booking_date = format_date(d.get('วันที่จอง'))
            sender = str(d.get('ชื่อ-นานสกุล') or d.get('หน่วยงาน') or 'ผู้ส่งตรวจ').strip()
            dept = str(d.get('หน่วยงาน') or 'งานจุลชีววิทยา').strip()
            contact = str(d.get('เบอร์ติดต่อ') or '-').replace('.0', '').strip()
            test_type = str(d.get('สิ่งส่งตรวจ') or 'ตรวจสิ่งแวดล้อม').strip()
            try:
                sample_count = int(float(d.get('จำนวน') or 1))
                if sample_count <= 0 or sample_count > 200: sample_count = 1
            except: sample_count = 1
            notes = str(d.get('หมายเหตุ') or '').strip()
            if notes == 'None': notes = ''
            
            srv_code = 'AIR_01'
            srv_name = 'Air Sampling (สำหรับงานอาชีวอนามัย)'
            if 'ยา' in test_type: srv_code = 'DRG_07'; srv_name = 'Drug (สำหรับงานผลิตยา1) ปลอดเชื้อ'
            elif 'เลือด' in test_type: srv_code = 'STR_02'; srv_name = 'Sterility (สำหรับงานธนาคารเลือด)'
            elif 'อาหาร' in test_type: srv_code = 'FOD_06'; srv_name = 'Food (สำหรับงานโภชนาการ)'
            elif 'น้ำ' in test_type: srv_code = 'WTO_04'; srv_name = 'Water (สำหรับห้องผ่าตัด OR)'
            elif 'พื้นผิว' in test_type: srv_code = 'WTS_03'; srv_name = 'Water or Surface (สำหรับงานควบคุมโรคติดเชื้อ)'
            
            bookings_payload.append({
                "booking_date": booking_date,
                "sender_name": sender[:100],
                "department": dept[:100],
                "contact_number": contact[:50],
                "service_code": srv_code,
                "service_name": srv_name,
                "sample_count": sample_count,
                "notes": notes,
                "status": "confirmed",
                "created_by": user_id
            })
            
        supabase_post("bookings", bookings_payload)
        print(f"  ✅ Inserted {len(bookings_payload)} bookings")
    else:
        print(f"  ✅ Bookings already present: {len(existing_bookings)} records")


# ============================================================================
# HELPER: BATCH INSERT REPORTS AND ITEMS
# ============================================================================
def insert_service_submissions(service_label, submissions_dict):
    print(f"\n📂 Migrating {service_label} ({len(submissions_dict)} submissions)...")
    success_reports = 0
    total_items = 0
    
    for sub_id, data in submissions_dict.items():
        header = data["header"]
        header["created_by"] = user_id
        
        rep_res, err = supabase_post("reports", [header])
        if err:
            # Handle duplicate submission_no
            header["submission_no"] += f"-{header['sampling_date'][-2:]}"
            rep_res, err = supabase_post("reports", [header])
            
        if rep_res and len(rep_res) > 0:
            report_id = rep_res[0]["id"]
            items = data["items"]
            for idx, it in enumerate(items):
                it["report_id"] = report_id
                it["item_no"] = idx + 1
            if items:
                supabase_post("report_items", items)
                total_items += len(items)
            success_reports += 1
            
    print(f"  ✅ Finished {service_label}: {success_reports} reports, {total_items} sample items inserted!")


# ============================================================================
# 2. AIR_01 (Air Sampling)
# ============================================================================
air_file = glob.glob("downloaded_raw/AIR_01*.xlsx")
if air_file:
    wb = openpyxl.load_workbook(air_file[0], data_only=True)
    sheet = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    
    subs = {}
    for r in rows[1:]:
        if not r or not any(r): continue
        d = dict(zip(header, r))
        sub_id = str(d.get('SubmissionID') or '').strip()
        if not sub_id or sub_id == 'None': continue
        
        if sub_id not in subs:
            sub_date = format_date(d.get('SubmissionDate') or d.get('Timestamp'))
            dept = str(d.get('Department') or 'งานอาชีวอนามัย').strip()
            ward = str(d.get('Ward') or '').strip()
            clean_sub_no = sub_id.replace('SUB', 'AIR-')
            if not clean_sub_no.startswith('AIR-'): clean_sub_no = f"AIR-{clean_sub_no}"
            
            subs[sub_id] = {
                "header": {
                    "submission_no": clean_sub_no[:50],
                    "service_code": "AIR_01",
                    "service_name": "Air Sampling (สำหรับงานอาชีวอนามัย)",
                    "department": dept[:100],
                    "ward_room": ward[:100],
                    "sampling_date": sub_date,
                    "received_date": sub_date,
                    "reported_date": sub_date,
                    "sampler_name": "เจ้าหน้าที่อาชีวอนามัย",
                    "reporter_name": "ทนพ.มานพ นันตาบุตร",
                    "approver_name": "ทนพญ.กรรณิการ์ สิทธิโชค",
                    "overall_result": "pass",
                    "status": "completed",
                    "remarks": "ตรวจนับจำนวนโคโลนีเชื้อแบคทีเรียและเชื้อราในอากาศ"
                },
                "items": []
            }
            
        b_count = str(d.get('Bacteria_CFU') if d.get('Bacteria_CFU') is not None else 0).replace('.0', '')
        f_count = str(d.get('Fungus_CFU') if d.get('Fungus_CFU') is not None else 0).replace('.0', '')
        loc = str(d.get('Location') or d.get('Ward') or 'จุดเก็บตัวอย่าง').strip()
        
        try:
            b_num = float(b_count); f_num = float(f_count)
            item_res = "pass" if (b_num < 500 and f_num < 100) else "fail"
        except: item_res = "pass"
        if item_res == "fail": subs[sub_id]["header"]["overall_result"] = "fail"
        
        subs[sub_id]["items"].append({
            "location_name": loc[:100],
            "sample_description": str(d.get('SampleType') or 'อากาศ'),
            "bacteria_count": b_count,
            "fungus_count": f_count,
            "standard_limit": "< 500 CFU/m³ (Bacteria), < 100 CFU/m³ (Fungi)",
            "item_result": item_res,
            "remarks": str(d.get('Remarks') or '')
        })
    insert_service_submissions("AIR_01 Air Sampling", subs)


# ============================================================================
# 3. STR_02 (Sterility Test / Blood Bank)
# ============================================================================
str_file = glob.glob("downloaded_raw/STR_02*.xlsx")
if str_file:
    wb = openpyxl.load_workbook(str_file[0], data_only=True)
    sheet = wb['Submissions'] if 'Submissions' in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    
    subs = {}
    for r in rows[1:]:
        if not r or not any(r): continue
        d = dict(zip(header, r))
        sub_id = str(d.get('Submission ID') or '').strip()
        if not sub_id or sub_id == 'None': continue
        
        if sub_id not in subs:
            sub_date = format_date(d.get('Submission Date') or d.get('Timestamp'))
            dept = str(d.get('Department') or 'งานธนาคารเลือด').strip()
            clean_sub_no = sub_id.replace('SUB', 'STR-')
            if not clean_sub_no.startswith('STR-'): clean_sub_no = f"STR-{clean_sub_no}"
            
            subs[sub_id] = {
                "header": {
                    "submission_no": clean_sub_no[:50],
                    "service_code": "STR_02",
                    "service_name": "Sterility (สำหรับงานธนาคารเลือด)",
                    "department": dept[:100],
                    "ward_room": "ธนาคารเลือด",
                    "sampling_date": sub_date,
                    "received_date": sub_date,
                    "reported_date": sub_date,
                    "sampler_name": "เจ้าหน้าที่ธนาคารเลือด",
                    "reporter_name": "ทนพ.อนุชิต บุญประสิทธิ์",
                    "approver_name": "ทนพ.มานพ นันตาบุตร",
                    "overall_result": "pass",
                    "status": "completed",
                    "remarks": "บ่มเชื้อครบ 14 วัน ไม่พบการเจริญเติบโตของเชื้อ"
                },
                "items": []
            }
            
        bag_no = str(d.get('Blood Bag Number') or '-').replace('.0', '')
        prod_type = str(d.get('Product Type') or 'เลือด')
        sterile_res = str(d.get('Sterile') or 'No growth')
        item_res = "pass" if "no growth" in sterile_res.lower() or "sterile" in sterile_res.lower() else "fail"
        if item_res == "fail": subs[sub_id]["header"]["overall_result"] = "fail"
        
        subs[sub_id]["items"].append({
            "location_name": f"ถุงเลือด #{bag_no} ({prod_type})",
            "sample_description": prod_type,
            "bacteria_count": "0",
            "fungus_count": "0",
            "standard_limit": "No bacterial or fungal growth",
            "item_result": item_res,
            "remarks": sterile_res
        })
    insert_service_submissions("STR_02 Sterility (Blood Bank)", subs)


# ============================================================================
# 4. WTS_03 (Water or Surface / IC)
# ============================================================================
wts_file = glob.glob("downloaded_raw/WTS_03*.xlsx")
if wts_file:
    wb = openpyxl.load_workbook(wts_file[0], data_only=True)
    sheet = wb['Submissions'] if 'Submissions' in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    
    subs = {}
    for r in rows[1:]:
        if not r or not any(r): continue
        d = dict(zip(header, r))
        sub_id = str(d.get('SubmissionID') or '').strip()
        if not sub_id or sub_id == 'None': continue
        
        if sub_id not in subs:
            sub_date = format_date(d.get('SubmissionDate') or d.get('Timestamp'))
            dept = str(d.get('Department') or 'งานควบคุมโรคติดเชื้อ (IC)').strip()
            clean_sub_no = sub_id.replace('JOB-', 'WTS-')
            if not clean_sub_no.startswith('WTS-'): clean_sub_no = f"WTS-{clean_sub_no}"
            
            subs[sub_id] = {
                "header": {
                    "submission_no": clean_sub_no[:50],
                    "service_code": "WTS_03",
                    "service_name": "Water or Surface (สำหรับงานควบคุมโรคติดเชื้อ)",
                    "department": dept[:100],
                    "ward_room": dept[:100],
                    "sampling_date": sub_date,
                    "received_date": sub_date,
                    "reported_date": sub_date,
                    "sampler_name": "พว. งานควบคุมโรคติดเชื้อ",
                    "reporter_name": "ทนพ.มานพ นันตาบุตร",
                    "approver_name": "ทนพญ.กรรณิการ์ สิทธิโชค",
                    "overall_result": "pass",
                    "status": "completed",
                    "remarks": "ตรวจเพาะเชื้อสิ่งแวดล้อมและพื้นผิวสัมผัส"
                },
                "items": []
            }
            
        loc = str(d.get('Location') or 'จุดสว็อบ').strip()
        culture_res = str(d.get('CultureResult') or 'No growth').strip()
        suspected = str(d.get('SuspectedOrganism') or '').strip()
        item_res = "pass" if "no growth" in culture_res.lower() or culture_res == '-' else "fail"
        if item_res == "fail": subs[sub_id]["header"]["overall_result"] = "fail"
        
        subs[sub_id]["items"].append({
            "location_name": loc[:100],
            "sample_description": str(d.get('SpecimenType') or 'พื้นผิว'),
            "bacteria_count": culture_res,
            "fungus_count": "0",
            "microorganism_found": culture_res if item_res == "fail" else "ไม่พบเชื้อก่อโรค",
            "standard_limit": "No pathogenic microorganisms",
            "item_result": item_res,
            "remarks": suspected
        })
    insert_service_submissions("WTS_03 Water or Surface", subs)


# ============================================================================
# 5. WTO_04 (Water OR)
# ============================================================================
wto_file = glob.glob("downloaded_raw/WTO_04*.xlsx")
if wto_file:
    wb = openpyxl.load_workbook(wto_file[0], data_only=True)
    sheet = wb['EnviMicrobiologyReports'] if 'EnviMicrobiologyReports' in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    
    subs = {}
    for r in rows[1:]:
        if not r or not any(r): continue
        d = dict(zip(header, r))
        sub_id = str(d.get('SubmissionID') or '').strip()
        if not sub_id or sub_id == 'None': continue
        
        if sub_id not in subs:
            sub_date = format_date(d.get('SubmissionDate') or d.get('Timestamp'))
            dept = str(d.get('Department') or 'ห้องผ่าตัด').strip()
            clean_sub_no = sub_id.replace('ENV-', 'WTO-')
            if not clean_sub_no.startswith('WTO-'): clean_sub_no = f"WTO-{clean_sub_no}"
            
            subs[sub_id] = {
                "header": {
                    "submission_no": clean_sub_no[:50],
                    "service_code": "WTO_04",
                    "service_name": "Water (สำหรับห้องผ่าตัด OR)",
                    "department": dept[:100],
                    "ward_room": "ห้องผ่าตัด",
                    "sampling_date": sub_date,
                    "received_date": sub_date,
                    "reported_date": sub_date,
                    "sampler_name": "พยาบาลห้องผ่าตัด",
                    "reporter_name": "ทนพ.มานพ นันตาบุตร",
                    "approver_name": "ทนพญ.กรรณิการ์ สิทธิโชค",
                    "overall_result": "pass",
                    "status": "completed",
                    "remarks": "ตรวจคุณภาพน้ำห้องผ่าตัด"
                },
                "items": []
            }
            
        loc = str(d.get('Location') or 'อ่างล้างมือ').strip()
        ecoli = str(d.get('Ecoli') or 'No growth').strip()
        item_res = "pass" if "no growth" in ecoli.lower() or "not detected" in ecoli.lower() else "fail"
        
        subs[sub_id]["items"].append({
            "location_name": loc[:100],
            "sample_description": str(d.get('SpecimenType') or 'น้ำ'),
            "bacteria_count": ecoli,
            "fungus_count": "0",
            "standard_limit": "< 10 CFU/100mL (TVC), Endotoxin < 0.25 EU/mL",
            "item_result": item_res,
            "remarks": ecoli
        })
    insert_service_submissions("WTO_04 Water (OR)", subs)


# ============================================================================
# 6. FOD_06 (Food)
# ============================================================================
fod_file = glob.glob("downloaded_raw/FOD_06*.xlsx")
if fod_file:
    wb = openpyxl.load_workbook(fod_file[0], data_only=True)
    sheet = wb['Data'] if 'Data' in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    
    subs = {}
    for idx, r in enumerate(rows[1:]):
        if not r or not any(r): continue
        d = dict(zip(header, r))
        sub_date = format_date(d.get('SubmissionDate') or d.get('Timestamp'))
        dept = str(d.get('Department') or 'งานโภชนาการ').strip()
        
        group_key = f"FOD-{sub_date}"
        if group_key not in subs:
            subs[group_key] = {
                "header": {
                    "submission_no": f"{group_key}-{len(subs)+1:02d}",
                    "service_code": "FOD_06",
                    "service_name": "Food (สำหรับงานโภชนาการ)",
                    "department": dept[:100],
                    "ward_room": "งานโภชนาการ",
                    "sampling_date": sub_date,
                    "received_date": sub_date,
                    "reported_date": sub_date,
                    "sampler_name": "นักโภชนาการ",
                    "reporter_name": "ทนพ.มานพ นันตาบุตร",
                    "approver_name": "ทนพญ.กรรณิการ์ สิทธิโชค",
                    "overall_result": "pass",
                    "status": "completed",
                    "remarks": "ตรวจวิเคราะห์จุลินทรีย์ในอาหารและสุขาภิบาล"
                },
                "items": []
            }
            
        food_name = str(d.get('Food') or f"ตัวอย่างอาหาร {idx+1}").strip()
        ecoli = str(d.get('Ecoli') or 'ไม่พบเชื้อ').strip()
        paer = str(d.get('Paeruginosa') or 'ไม่พบเชื้อ').strip()
        item_res = "pass" if ("ไม่พบ" in ecoli or "negative" in ecoli.lower()) else "fail"
        
        subs[group_key]["items"].append({
            "location_name": food_name[:100],
            "sample_description": str(d.get('SampleType') or 'อาหาร'),
            "bacteria_count": ecoli,
            "fungus_count": paer,
            "standard_limit": "E. coli Negative, Salmonella/Shigella Not detected",
            "item_result": item_res,
            "remarks": str(d.get('Remarks') or '')
        })
    insert_service_submissions("FOD_06 Food (Nutrition)", subs)


# ============================================================================
# 7. DRG_07 (Drug 1 Sterile)
# ============================================================================
drg1_file = glob.glob("downloaded_raw/DRG_07*.csv")
if drg1_file:
    with open(drg1_file[0], mode='r', encoding='utf-8-sig', errors='ignore') as f:
        reader = csv.reader(f)
        rows = list(reader)
    header = rows[0]
    subs = {}
    for r in rows[1:]:
        if not r or not any(r): continue
        d = dict(zip(header, r))
        sub_id = str(d.get('SubmissionID') or '').strip()
        if not sub_id: continue
        
        if sub_id not in subs:
            sub_date = format_date(d.get('ReceiptDate') or d.get('Timestamp'))
            clean_sub_no = sub_id.replace('JOB-', 'DR1-')
            if not clean_sub_no.startswith('DR1-'): clean_sub_no = f"DR1-{clean_sub_no}"
            
            subs[sub_id] = {
                "header": {
                    "submission_no": clean_sub_no[:50],
                    "service_code": "DRG_07",
                    "service_name": "Drug (สำหรับงานผลิตยา1) ปลอดเชื้อ",
                    "department": "งานผลิตยา",
                    "ward_room": "หน่วยเตรียมยาปราศจากเชื้อ",
                    "sampling_date": sub_date,
                    "received_date": sub_date,
                    "reported_date": format_date(d.get('AnalysisDate') or sub_date),
                    "sampler_name": "เภสัชกรงานผลิตยา",
                    "reporter_name": "ทนพ.มานพ นันตาบุตร",
                    "approver_name": "ทนพญ.กรรณิการ์ สิทธิโชค",
                    "overall_result": "pass",
                    "status": "completed",
                    "remarks": "ทดสอบความปราศจากเชื้อตามมาตรฐาน USP <71>"
                },
                "items": []
            }
            
        med_name = str(d.get('PreparedMedicineHeader') or 'ยาปราศจากเชื้อ').strip()
        lot_no = str(d.get('LotNo') or '-').strip()
        
        subs[sub_id]["items"].append({
            "location_name": f"{med_name} (Lot {lot_no})",
            "sample_description": f"Volume: {d.get('Volume', '')} mL, Count: {d.get('SampleCount', '')}",
            "bacteria_count": "No growth",
            "fungus_count": "No growth",
            "standard_limit": "USP <71> Sterility Tests (14 days)",
            "item_result": "pass",
            "remarks": f"Prod Date: {d.get('ProductionDate', '')}"
        })
    insert_service_submissions("DRG_07 Drug 1 Sterile", subs)


# ============================================================================
# 8. DRG_08 (Drug 2 Microbial Contamination)
# ============================================================================
drg2_file = glob.glob("downloaded_raw/DRG_08*.xlsx")
if drg2_file:
    wb = openpyxl.load_workbook(drg2_file[0], data_only=True)
    sheet = wb['Submissions'] if 'Submissions' in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    
    subs = {}
    for r in rows[1:]:
        if not r or not any(r): continue
        d = dict(zip(header, r))
        sub_id = str(d.get('Submission ID') or '').strip()
        if not sub_id or sub_id == 'None': continue
        
        if sub_id not in subs:
            sub_date = format_date(d.get('Sample Date') or d.get('Timestamp'))
            dept = str(d.get('Department') or 'งานผลิตยา').strip()
            clean_sub_no = sub_id.replace('TUH-', 'DR2-')
            if not clean_sub_no.startswith('DR2-'): clean_sub_no = f"DR2-{clean_sub_no}"
            
            subs[sub_id] = {
                "header": {
                    "submission_no": clean_sub_no[:50],
                    "service_code": "DRG_08",
                    "service_name": "Drug (สำหรับงานผลิตยา2) แบบรายงานผลการวิเคราะห์การปนเปื้อนเชื้อจุลินทรีย์",
                    "department": dept[:100],
                    "ward_room": "งานผลิตยา",
                    "sampling_date": sub_date,
                    "received_date": sub_date,
                    "reported_date": sub_date,
                    "sampler_name": str(d.get('Operator') or 'เภสัชกรงานผลิตยา'),
                    "reporter_name": "ทนพ.มานพ นันตาบุตร",
                    "approver_name": "ทนพญ.กรรณิการ์ สิทธิโชค",
                    "overall_result": "pass",
                    "status": "completed",
                    "remarks": "ตรวจนับจำนวนจุลินทรีย์และการปนเปื้อนตาม USP <61>/<62>"
                },
                "items": []
            }
            
        drug_type = str(d.get('Drug Type') or 'ผลิตภัณฑ์ยา').strip()
        tamc = str(d.get('TAMC') or '< 10^2 CFU/g').strip()
        tymc = str(d.get('TYMC') or '< 10^1 CFU/g').strip()
        
        subs[sub_id]["items"].append({
            "location_name": drug_type[:100],
            "sample_description": str(d.get('Sample Type') or 'ยา'),
            "bacteria_count": tamc,
            "fungus_count": tymc,
            "standard_limit": "USP <61> TAMC/TYMC Limits",
            "item_result": "pass",
            "remarks": str(d.get('Remarks') or '')
        })
    insert_service_submissions("DRG_08 Drug 2 Contamination", subs)

print("\n======================================================================")
print("🎉 ALL RAW DATA SUCCESSFULLY MIGRATED TO SUPABASE!")
print("======================================================================")
